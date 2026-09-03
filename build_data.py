#!/usr/bin/env python3
"""
build_data.py

Reads the weekly AMP AIM sample-tracking export (headers on row 3) and produces
a fully AGGREGATE, de-identified dashboard.json for the public progress dashboard.

Privacy design (do not change without re-checking with the data owner):
  - No Subject_ID, Subject-Visit_ID, or any other row-level identifier is ever
    written to the output.
  - No diagnosis (Enroll_Dx), demographic field (Sex/Race/Ethnicity/Age), or
    free-text Notes value is ever written to the output.
  - Only COUNTS are emitted, grouped by technology / disease-tissue scope /
    pipeline / cohort. Cells with fewer than MIN_CELL_COUNT contributing rows
    are still reported (counts, not identities) but the script makes it easy
    to raise that floor if small-cell suppression is ever required.

Usage:
    python3 build_data.py [input.xlsx] [output.json] [expected_recruitment.xlsx]

Defaults: data/AMP_AIM_Dataset.xlsx -> data/dashboard.json, reading recruitment
targets from data/Expected_Recruitment_Numbers.xlsx.
"""
import sys
import json
import datetime
from collections import defaultdict

import openpyxl

HEADER_ROW = 3
DATA_START_ROW = 4

# Technology (sample/assay) columns -> (key, display label)
# Trimmed to the 12 technologies actually tracked on the dashboard. (Plain
# Xenium, both "-NL" / Non-lesional columns, and the old OMRF Lab / Xenium
# (Slide) labels were dropped or renamed - see README history / requests.)
TECH_COLUMNS = [
    ("Xenium-Slide_Sample_ID", "xenium_slide", "Xenium"),
    ("scFFPE_Sample_ID", "scffpe", "scFFPE"),
    ("scRNAseq_Sample_ID", "scrnaseq", "scRNA-seq"),
    ("Olink-Serum_Sample_ID", "olink_serum", "Olink (Serum)"),
    ("CyTOF-Blood_Sample_ID", "cytof_blood", "CyTOF (Blood)"),
    ("Genotyping_Sample_ID", "genotyping", "Genotyping"),
    ("OMRF-Lab_Sample_ID", "omrf_lab", "Autoantibody Testing"),
    ("bRNAseq_Sample_ID", "brnaseq", "Bulk RNA-Seq"),
    ("Olink-Urine_Sample_ID", "olink_urine", "Olink (Urine)"),
    ("mBioSeq-stool", "mbioseq_stool", "Microbiome (Stool)"),
    ("mBioSeq-skin", "mbioseq_skin", "Microbiome (Skin)"),
    ("anti-C1Q_Sample_ID", "anti_c1q", "Anti-C1Q"),
]

STATUS_ORDER = ["completed", "pending", "not_applicable", "qc_fail", "unknown"]
STATUS_LABELS = {
    "completed": "Completed",
    "pending": "Pending / Specimen Available",
    "not_applicable": "Not Applicable / No Specimen",
    "qc_fail": "QC Fail",
    "unknown": "Unknown / Not Yet Retrieved",
}

# Visit codes that mark a subject's ENROLLMENT into a cohort (one per Visit_Type:
# Scheduled->V01, Control->VC1, Enabling->VE1, Archival->VA1, Unscheduled->VU1).
# The Visit_Cohort tag(s) on this row are the subject's cohort assignment(s).
# Checked in this priority order when a subject has more than one such row.
ENROLLMENT_CODES_PRIORITY = ["V01", "VC1", "VE1", "VA1", "VU1"]

# "Disease Team" in the source is unreliable (mostly #REF! - a broken lookup
# formula upstream). Data_Scope ("<disease>-<tissue>", e.g. "SLE-KDY") is
# populated and reliable, so the disease/team grouping is derived from its
# prefix instead. EDIT THIS if your actual disease-team names differ, or add
# an explicit Subject_ID -> team override if Data_Scope isn't authoritative
# for some subjects.
DISEASE_LABELS = {
    "SLE": "Lupus (SLE)",
    "RA": "Rheumatoid Arthritis (RA)",
    "PsD": "Psoriatic Disease (PsD)",
    "SjD": "Sjögren's Disease (SjD)",
    "SSc": "Systemic Sclerosis (SSc)",
    "SLE/PsD": "SLE / PsD (combined)",
}

# SLE and PsD are further split by tissue (the part of Data_Scope after the
# hyphen) into the finer subgroups the dashboard now reports as their own
# "disease teams". (prefix, tissue) -> (key, display label). Any prefix/tissue
# combo NOT listed here just falls back to the whole-prefix grouping above
# (e.g. RA-SYN, SjD-SGL, SSc-SKN all stay as one bucket per prefix).
SPLIT_DISEASE_LABELS = {
    ("SLE", "KDY"): ("sle_kdy", "Lupus Kidney"),
    ("SLE", "SKN"): ("sle_skn", "Lupus Skin"),
    ("PsD", "SYN"): ("psd_syn", "Psoriatic Arthritis"),
    ("PsD", "SKN"): ("psd_skn", "Psoriasis"),
    ("PsD", "EYE"): ("psd_eye", "Uveitis"),
}

# Label used for a blank/None Pipeline cell.
UNDEFINED_PIPELINE = "Undefined"

# A subject's enrollment Visit_Code also tells us their recruitment status:
# VA1 rows are archival specimens, everything else is a normal enrolled visit.
STATUS_BY_CODE = {
    "V01": "enrolled",
    "VC1": "enrolled",
    "VE1": "enrolled",
    "VA1": "archival",
    "VU1": "enrolled",
}
RECRUIT_STATUS_ORDER = ["enrolled", "archival", "unknown"]
RECRUIT_STATUS_LABELS = {"enrolled": "Enrolled", "archival": "Archival", "unknown": "Unknown"}

# How many subjects are ultimately expected in each cohort, per the network's
# recruitment targets. Loaded at runtime (see load_expected_recruitment(),
# called from main()) from a spreadsheet - data/Expected_Recruitment_Numbers.xlsx
# by default - with columns "Disease Team", "Cohort", "Expected" (an optional
# "Notes" column is ignored). Edit that spreadsheet directly to update targets;
# no code change needed. A blank or "Undefined" Expected cell means "not yet
# defined". Keyed by the ORIGINAL (pre-split) disease-team label, since that's
# the level the targets were given at; a cohort's new, finer disease-subgroup
# membership (see SPLIT_DISEASE_LABELS) is looked up from the real data, and
# this table is then consulted by that cohort's old/whole-disease label.
EXPECTED_RECRUITMENT_PATH_DEFAULT = "data/Expected_Recruitment_Numbers.xlsx"
EXPECTED_RECRUITMENT = {}  # populated by load_expected_recruitment() in main()


def load_expected_recruitment(path):
    """
    Read the recruitment-target spreadsheet and return
    {disease_team_label: {cohort_name_lowercased: int_or_None}}.
    Returns {} (every target "not yet set") if the file is missing or doesn't
    have the expected headers - the dashboard still builds fine either way,
    it just won't show "of N" progress numbers until the file is in place.
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except FileNotFoundError:
        print(f"WARNING: recruitment-target file not found at {path!r} - "
              f"recruitment bars will show as \"not yet set\" until it's added.",
              file=sys.stderr)
        return {}
    ws = wb[wb.sheetnames[0]]
    headers = [c.value.strip() if isinstance(c.value, str) else c.value for c in ws[1]]
    col_idx = {h: i for i, h in enumerate(headers) if h}
    required = ["Disease Team", "Cohort", "Expected"]
    missing = [c for c in required if c not in col_idx]
    if missing:
        print(f"WARNING: {path!r} is missing column(s) {missing} (expected "
              f"'Disease Team', 'Cohort', 'Expected' headers on row 1) - "
              f"recruitment bars will show as \"not yet set\".", file=sys.stderr)
        return {}

    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        team_cell = row[col_idx["Disease Team"]]
        cohort_cell = row[col_idx["Cohort"]]
        if team_cell is None or cohort_cell is None:
            continue
        team = str(team_cell).strip()
        cohort = str(cohort_cell).strip().lower()
        raw = row[col_idx["Expected"]]
        value = int(raw) if isinstance(raw, (int, float)) else None
        out.setdefault(team, {})[cohort] = value
    return out


def expected_for(old_label, cohort_name):
    """Look up a cohort's recruitment target. None if not defined/found."""
    return EXPECTED_RECRUITMENT.get(old_label, {}).get(cohort_name.strip().lower())


def empty_recruit_status():
    return {k: 0 for k in RECRUIT_STATUS_ORDER}


def derive_diseases(scope_label):
    """
    Return the list of (key, label, old_label, old_group_key) disease-team
    entries a given Data_Scope value belongs to. Usually a single entry (e.g.
    'RA-SYN' -> [('ra', 'Rheumatoid Arthritis (RA)', 'Rheumatoid Arthritis
    (RA)', 'ra')]), but Data_Scope can encode more than one disease/tissue in
    a single cell with a '/' - either in the prefix ('SLE/PsD-SKN', an
    overlap-syndrome subject) or in the tissue ('PsD-SKN/SYN', a subject with
    data under both tissues). In either case the subject is counted in EACH
    matching disease team, but only once in the site-wide subject total (that
    total is a distinct count of Subject_ID and never goes through this
    function).
    `old_label`/`old_group_key` are the pre-split whole-disease label/key
    (e.g. still "Lupus (SLE)" / "sle" for both Lupus subgroups) - used to
    roll the new finer subgroups back up into one combined bar with one
    combined Expected Recruitment target, since that's the level those
    targets were given at.
    """
    if not scope_label or scope_label == "Unknown":
        return [("unknown", "Unknown", "Unknown", "unknown")]
    if "-" in scope_label:
        prefix_part, tissue_part = scope_label.split("-", 1)
    else:
        prefix_part, tissue_part = scope_label, ""
    prefixes = [p.strip() for p in prefix_part.split("/") if p.strip()] or [prefix_part.strip()]
    tissues = [t.strip() for t in tissue_part.split("/") if t.strip()] or [""]

    seen = {}
    for prefix in prefixes:
        old_label = DISEASE_LABELS.get(prefix, prefix)
        old_group_key = prefix.lower().replace(" ", "_")
        for tissue in tissues:
            split = SPLIT_DISEASE_LABELS.get((prefix, tissue))
            if split:
                key, label = split
            else:
                key = old_group_key
                label = old_label
            if key not in seen:
                seen[key] = (key, label, old_label, old_group_key)
    return list(seen.values())


def classify(value):
    """Collapse the many raw status strings into 5 reportable buckets."""
    if value is None:
        return "unknown"
    s = str(value).strip()
    if s == "" or s in ("None", "#REF!", "[not yet retrieved]"):
        return "unknown"
    low = s.lower()
    if "qc fail" in low or "qc-fail" in low or "failed qc" in low:
        return "qc_fail"
    if "not applicable" in low or "no specimen" in low or "not available" in low:
        return "not_applicable"
    if "not ordered" in low:
        return "not_applicable"
    if "specimen available" in low or "pending" in low:
        return "pending"
    if s.startswith("[") and s.endswith("]"):
        # any other bracketed status we haven't explicitly mapped
        return "unknown"
    # anything else is an actual sample/barcode ID (or "Yes") -> completed
    return "completed"


def clean_label(value, default="Unknown"):
    """Clean a single-value field (Data_Scope, Pipeline, ...) into one label."""
    if value is None:
        return default
    s = str(value).strip()
    if s == "" or s in ("None", "#REF!"):
        return default
    if s.startswith("[") and s.endswith("]"):
        inner = s[1:-1]
        parts = [p.strip() for p in inner.split("][") if p.strip()]
        s = ", ".join(parts) if parts else default
    if s.lower() == "not yet retrieved":
        s = "Not yet retrieved"
    return s


def split_cohort_tags(value):
    """
    Visit_Cohort can hold MULTIPLE tags back-to-back in one cell, e.g.
    "[Cohort 1a: PsO][Cohort 1b: Drug-Naive PsA]" for a subject who belongs to
    both. Returns a LIST of individual cohort tags rather than one combined
    string, so a subject like that is counted once in EACH cohort - never as
    a separate "Cohort 1a, Cohort 1b" bucket.
    """
    if value is None:
        return ["Unknown"]
    s = str(value).strip()
    if s == "" or s in ("None", "#REF!"):
        return ["Unknown"]
    if s.startswith("[") and s.endswith("]"):
        inner = s[1:-1]
        parts = [p.strip() for p in inner.split("][") if p.strip()]
        if not parts:
            return ["Unknown"]
        return ["Not yet retrieved" if p.lower() == "not yet retrieved" else p for p in parts]
    if s.lower() == "not yet retrieved":
        return ["Not yet retrieved"]
    return [s]


def empty_status_counts():
    return {k: 0 for k in STATUS_ORDER}


def add_counts(dst, src):
    for k in STATUS_ORDER:
        dst[k] = dst.get(k, 0) + (src.get(k, 0) if src else 0)
    return dst


def main():
    in_path = sys.argv[1] if len(sys.argv) > 1 else "data/AMP_AIM_Dataset.xlsx"
    out_path = sys.argv[2] if len(sys.argv) > 2 else "data/dashboard.json"
    expected_path = sys.argv[3] if len(sys.argv) > 3 else EXPECTED_RECRUITMENT_PATH_DEFAULT

    global EXPECTED_RECRUITMENT
    EXPECTED_RECRUITMENT = load_expected_recruitment(expected_path)

    wb = openpyxl.load_workbook(in_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [ws.cell(row=HEADER_ROW, column=c).value for c in range(1, ws.max_column + 1)]
    col_idx = {h: i + 1 for i, h in enumerate(headers) if h}

    missing = [col for col, _, _ in TECH_COLUMNS if col not in col_idx]
    if missing:
        print(f"WARNING: expected columns not found and will be skipped: {missing}", file=sys.stderr)

    subject_idx = col_idx.get("Subject_ID")
    data_scope_idx = col_idx.get("Data_Scope")
    visit_cohort_idx = col_idx.get("Visit_Cohort")
    visit_code_idx = col_idx.get("Visit_Code")
    pipeline_idx = col_idx.get("Pipeline")

    subjects_seen = set()
    n_visits = 0
    # subject -> list of (visit_code_raw, cohort_cell_raw_value, scope_label)
    # cohort_cell_raw_value is kept RAW (not yet split) so pick_enrollment()
    # below can split it into individual tags itself.
    subj_rows = defaultdict(list)

    # tech key -> overall status counts
    tech_totals = {key: empty_status_counts() for _, key, _ in TECH_COLUMNS}
    # tech key -> {scope_label: status_counts}  (all pipelines combined)
    tech_by_scope = {key: defaultdict(empty_status_counts) for _, key, _ in TECH_COLUMNS}
    # tech key -> {pipeline_label: status_counts}  (all scopes combined)
    tech_by_pipeline = {key: defaultdict(empty_status_counts) for _, key, _ in TECH_COLUMNS}
    # tech key -> {scope_label: {pipeline_label: status_counts}}  (for disease-filtered pipeline views)
    tech_by_scope_pipeline = {key: defaultdict(lambda: defaultdict(empty_status_counts)) for _, key, _ in TECH_COLUMNS}

    scopes_seen = set()
    pipelines_seen = set()

    for r in range(DATA_START_ROW, ws.max_row + 1):
        subj = ws.cell(row=r, column=subject_idx).value if subject_idx else None
        subj_clean = clean_label(subj, default=None)
        # skip fully blank rows
        row_has_any = any(
            ws.cell(row=r, column=col_idx[col]).value is not None
            for col, _, _ in TECH_COLUMNS
            if col in col_idx
        )
        if subj_clean is None and not row_has_any:
            continue

        if subj_clean:
            subjects_seen.add(subj_clean)
        n_visits += 1

        scope_label = clean_label(ws.cell(row=r, column=data_scope_idx).value if data_scope_idx else None)
        pipeline_label = clean_label(
            ws.cell(row=r, column=pipeline_idx).value if pipeline_idx else None,
            default=UNDEFINED_PIPELINE,
        )
        scopes_seen.add(scope_label)
        pipelines_seen.add(pipeline_label)

        if subj_clean:
            code_raw = ws.cell(row=r, column=visit_code_idx).value if visit_code_idx else None
            code_raw = str(code_raw).strip() if code_raw is not None else ""
            cohort_raw = ws.cell(row=r, column=visit_cohort_idx).value if visit_cohort_idx else None
            subj_rows[subj_clean].append((code_raw, cohort_raw, scope_label))

        for col, key, _ in TECH_COLUMNS:
            if col not in col_idx:
                continue
            val = ws.cell(row=r, column=col_idx[col]).value
            status = classify(val)
            tech_totals[key][status] += 1
            tech_by_scope[key][scope_label][status] += 1
            tech_by_pipeline[key][pipeline_label][status] += 1
            tech_by_scope_pipeline[key][scope_label][pipeline_label][status] += 1

    def totals_block(counts):
        total = sum(counts.values())
        completed = counts["completed"]
        pct = round(100 * completed / total, 1) if total else 0.0
        eligible = counts["completed"] + counts["pending"] + counts["qc_fail"]
        pct_elig = round(100 * completed / eligible, 1) if eligible else None
        return {
            "counts": counts,
            "total": total,
            "pct_complete": pct,
            "eligible": eligible,
            "pct_eligible": pct_elig,
        }

    technologies = []
    for _, key, label in TECH_COLUMNS:
        block = totals_block(tech_totals[key])
        block["key"] = key
        block["label"] = label
        technologies.append(block)
    technologies.sort(key=lambda t: t["counts"]["completed"], reverse=True)

    scopes_sorted = sorted(s for s in scopes_seen if s != "Unknown") + (
        ["Unknown"] if "Unknown" in scopes_seen else []
    )
    pipelines_sorted = sorted(pipelines_seen)  # "EDP1","EDP2","NRP","Undefined" sorts naturally

    scope_disease = {}
    for scope in scopes_sorted:
        scope_disease[scope] = [{"key": k, "label": l} for k, l, _old, _grp in derive_diseases(scope)]

    # ---------- Recruitment: subjects & visits per disease team / cohort ----------
    def pick_enrollment(rows):
        """
        Pick the (cohort_tags, scope, code) for a subject from their
        enrollment-code row(s). cohort_tags is a LIST (a subject can be in
        more than one cohort at once). Prefers, in ENROLLMENT_CODES_PRIORITY
        order, a row that actually has a real (non-"Unknown") cohort tag.
        The code itself (V01/VC1/VE1/VA1/VU1) also tells us archival status.
        """
        by_code = {}
        for code, cohort_raw, scope in rows:
            if code not in ENROLLMENT_CODES_PRIORITY:
                continue
            tags = split_cohort_tags(cohort_raw)
            has_real = any(t != "Unknown" for t in tags)
            if code not in by_code or (not by_code[code][2] and has_real):
                by_code[code] = (tags, scope, has_real)
        for code in ENROLLMENT_CODES_PRIORITY:
            if code in by_code and by_code[code][2]:
                return by_code[code][0], by_code[code][1], code
        for code in ENROLLMENT_CODES_PRIORITY:
            if code in by_code:
                return by_code[code][0], by_code[code][1], code
        return None

    disease_totals = defaultdict(lambda: {
        "subjects": 0, "visits": 0, "label": "", "old_label": "", "old_group_key": "",
        "status": empty_recruit_status(),
    })
    cohort_totals = defaultdict(lambda: {
        "subjects": 0, "visits": 0, "disease_key": "", "disease_label": "", "old_label": "",
        "status": empty_recruit_status(),
    })
    # Rolled back up to the ORIGINAL, pre-split disease team (e.g. "sle"
    # combines sle_kdy + sle_skn) - this is where a single, shared Expected
    # Recruitment target still makes sense (see the "combined bar" note
    # above derive_diseases). Subject membership here is a real set, so a
    # subject touching both Lupus subgroups still counts once.
    combined_totals = defaultdict(lambda: {"subject_set": set(), "visits": 0, "label": "", "status": empty_recruit_status()})
    unassigned_subjects = 0

    for subj, rows in subj_rows.items():
        picked = pick_enrollment(rows)
        n_subject_visits = len(rows)
        if picked is None:
            unassigned_subjects += 1
            continue
        tags, scope, code = picked
        recruit_status = STATUS_BY_CODE.get(code, "unknown")
        touched_groups = {}
        # A subject can land in more than one disease-team bucket - see
        # derive_diseases() for the combo Data_Scope cases ("PsD-SKN/SYN",
        # "SLE/PsD-SKN", etc). They're counted once in EACH matching team,
        # but the site-wide subject total is a separate, plain distinct-
        # Subject_ID count (below) so it's never inflated by this.
        for disease_key, disease_label, old_label, old_group_key in derive_diseases(scope):
            disease_totals[disease_key]["subjects"] += 1
            disease_totals[disease_key]["visits"] += n_subject_visits
            disease_totals[disease_key]["label"] = disease_label
            disease_totals[disease_key]["old_label"] = old_label
            disease_totals[disease_key]["old_group_key"] = old_group_key
            disease_totals[disease_key]["status"][recruit_status] += 1
            touched_groups[old_group_key] = old_label
            # ...and within that team, counts in EVERY cohort tag they carry
            # (a subject in both "Cohort 1a" and "Cohort 1b" adds 1 to each).
            for tag in sorted(set(tags)):
                ck = (disease_key, tag)
                cohort_totals[ck]["subjects"] += 1
                cohort_totals[ck]["visits"] += n_subject_visits
                cohort_totals[ck]["disease_key"] = disease_key
                cohort_totals[ck]["disease_label"] = disease_label
                cohort_totals[ck]["old_label"] = old_label
                cohort_totals[ck]["status"][recruit_status] += 1
        # Roll this subject up into each ORIGINAL (pre-split) team they
        # touched - once each, even if they hit more than one new subgroup
        # within that same original team.
        for group_key, old_label in touched_groups.items():
            g = combined_totals[group_key]
            g["subject_set"].add(subj)
            g["visits"] += n_subject_visits
            g["label"] = old_label
            g["status"][recruit_status] += 1

    def disease_expected(old_label, cohort_names):
        """
        Sum this disease team's cohort-level recruitment targets. Returns
        (total_or_None, partial): partial=True means at least one
        contributing cohort has no target defined yet, so the sum is a
        floor (actual target is >= this), not the full picture.
        """
        total, any_known, any_unknown = 0, False, False
        for name in cohort_names:
            v = expected_for(old_label, name)
            if v is None:
                any_unknown = True
            else:
                total += v
                any_known = True
        return (total if any_known else None), any_unknown

    # Fine-grained (Lupus Kidney / Lupus Skin / ... ) - used for the disease
    # filter pills, KPI totals, and the per-subgroup cohort cards. No target
    # numbers at this granularity (see the module docstring / derive_diseases
    # note) - those live one level up, on by_disease_group below.
    by_disease = sorted(
        [{"key": k, "label": v["label"], "subjects": v["subjects"], "visits": v["visits"], "status": v["status"]}
         for k, v in disease_totals.items()],
        key=lambda d: d["subjects"], reverse=True,
    )

    # Rolled back up to the original disease team - one shared Expected
    # Recruitment target per team, with the fine-grained subgroups above
    # broken out as "segments" so a single progress-toward-target bar can
    # still show the Kidney/Skin (or Arthritis/Psoriasis/Uveitis) split.
    subgroup_to_group = {k: v["old_group_key"] for k, v in disease_totals.items()}
    by_disease_group = []
    for gk, g in combined_totals.items():
        member_keys = [k for k, og in subgroup_to_group.items() if og == gk]
        segments = sorted(
            [{"key": k, "label": disease_totals[k]["label"], "subjects": disease_totals[k]["subjects"]}
             for k in member_keys],
            key=lambda s: s["subjects"], reverse=True,
        )
        cohort_names = {ck[1] for ck, cv in cohort_totals.items() if cv["disease_key"] in member_keys}
        expected, partial = disease_expected(g["label"], cohort_names)
        by_disease_group.append({
            "key": gk, "label": g["label"], "subjects": len(g["subject_set"]), "visits": g["visits"],
            "status": g["status"], "expected": expected, "expected_partial": partial, "segments": segments,
        })
    by_disease_group.sort(key=lambda d: d["subjects"], reverse=True)

    by_cohort_detail = sorted(
        [{"disease_key": v["disease_key"], "disease_label": v["disease_label"], "cohort": ck[1],
          "subjects": v["subjects"], "visits": v["visits"], "status": v["status"],
          "expected": expected_for(v["old_label"], ck[1])}
         for ck, v in cohort_totals.items()],
        # Alphabetical by cohort name within each disease team.
        key=lambda d: (d["disease_label"], d["cohort"].lower()),
    )
    recruitment = {
        "by_disease": by_disease,
        "by_disease_group": by_disease_group,
        "by_cohort": by_cohort_detail,
        "unassigned_subjects": unassigned_subjects,
        "status_order": RECRUIT_STATUS_ORDER,
        "status_labels": RECRUIT_STATUS_LABELS,
    }

    by_scope_matrix = {
        key: {scope: tech_by_scope[key].get(scope, empty_status_counts()) for scope in scopes_sorted}
        for _, key, _ in TECH_COLUMNS
    }
    by_pipeline_matrix = {
        key: {p: tech_by_pipeline[key].get(p, empty_status_counts()) for p in pipelines_sorted}
        for _, key, _ in TECH_COLUMNS
    }
    by_scope_pipeline_matrix = {
        key: {
            scope: {p: tech_by_scope_pipeline[key][scope].get(p, empty_status_counts()) for p in pipelines_sorted}
            for scope in scopes_sorted
        }
        for _, key, _ in TECH_COLUMNS
    }

    output = {
        "generated_at": datetime.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source_rows": n_visits,
        "totals": {
            "subjects": len(subjects_seen),
            "visits": n_visits,
        },
        "recruitment": recruitment,
        "status_order": STATUS_ORDER,
        "status_labels": STATUS_LABELS,
        "technologies": technologies,
        "scopes": scopes_sorted,
        "pipelines": pipelines_sorted,
        "scope_disease": scope_disease,
        "by_scope": by_scope_matrix,
        "by_pipeline": by_pipeline_matrix,
        "by_scope_pipeline": by_scope_pipeline_matrix,
    }

    with open(out_path, "w") as f:
        json.dump(output, f, indent=2)

    print(f"Wrote {out_path}")
    print(f"  subjects={len(subjects_seen)} visits={n_visits} technologies={len(technologies)}")
    print(f"  disease teams={len(by_disease)} cohorts={len(by_cohort_detail)} unassigned_subjects={unassigned_subjects}")
    print(f"  pipelines={pipelines_sorted}")


if __name__ == "__main__":
    main()
